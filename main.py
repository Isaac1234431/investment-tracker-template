from fastapi import FastAPI, HTTPException, UploadFile, File
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Any
import base64
import io
import os
import re
import zipfile
import httpx
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.formula import ArrayFormula
from datetime import datetime
from collections import defaultdict

app = FastAPI()

# ── CORS — allows the UI (any origin) to call this service directly ──────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Constants ────────────────────────────────────────────────────────────────
TEMPLATE_URL = os.environ.get("TEMPLATE_URL", "")  # Set in Render env vars
N8N_WEBHOOK_URL = os.environ.get("N8N_WEBHOOK_URL", "")  # Set in Render env vars
DATE_FMT     = "mmm/dd/yyyy"
BLUE_FILL    = PatternFill("solid", fgColor="1F4E79")
WHITE_BOLD   = Font(bold=True, color="FFFFFF")
WHITE_FONT   = Font(color="FFFFFF")
PEND_COLS    = ["C","D","E","F","G","H","I","J","K","L","M","N","O","P"]

BUY_VARIANTS  = {"buy", "purchase", "bought"}
SELL_VARIANTS = {"sell", "sale", "sold", "redemption"}

# ── Request model ─────────────────────────────────────────────────────────────
class BuildRequest(BaseModel):
    extract_result: dict[str, Any]

# ── Helpers ───────────────────────────────────────────────────────────────────
def normalize_trans_type(raw: str) -> str:
    lower = raw.lower().strip()
    if lower in BUY_VARIANTS:  return "Buy"
    if lower in SELL_VARIANTS: return "Sell"
    return raw  # preserve original for non-trade types

def build_symbol_lookup(holdings):
    return {h["description"].upper().strip(): h["symbol"]
            for h in holdings if h.get("symbol") and h.get("description")}


def resolve_symbol(description, symbol, lookup):
    # 1. Use the symbol extracted directly from the transaction row
    if symbol:
        return symbol
    # 2. Exact description match against holdings
    return lookup.get((description or "").upper().strip(), "")


def parse_date(s):
    if not s: return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y"):
        try: return datetime.strptime(s, fmt)
        except: pass
    return s

def set_date(cell, value):
    cell.value = value
    if isinstance(value, datetime):
        cell.number_format = DATE_FMT

def clear_row(ws, row, cols):
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font()

def apply_blue_row(ws, row, cols, label=None, label_col=None):
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.fill = BLUE_FILL
        cell.font = WHITE_BOLD if (label and col == label_col) else WHITE_FONT
    if label and label_col:
        ws[f"{label_col}{row}"].value = label

def deduplicate_holdings(holdings):
    """
    Consolidate holdings that share the same ticker (or same description when no
    ticker exists) into a single row.  market_value, quantity, and cost fields are
    summed; all other fields are taken from the first lot seen.  This prevents
    duplicate ticker rows in the Summary sheet and ensures SUMPRODUCT formulas
    don't double-count transactions for the same security.
    """
    seen   = {}   # key -> consolidated holding dict
    order  = []   # preserve stable sort order (first occurrence wins)
    for h in holdings:
        key = (h.get("symbol") or "").upper().strip() or (h.get("description") or "").upper().strip()
        if not key:
            continue
        if key not in seen:
            seen[key] = {**h}   # shallow copy so we don't mutate the original
            order.append(key)
        else:
            # Accumulate numeric fields; leave descriptive fields as-is
            for field in ("market_value", "quantity"):
                existing = seen[key].get(field) or 0
                incoming = h.get(field) or 0
                seen[key][field] = existing + incoming
    return [seen[k] for k in order]

def compute_securities_value(holdings):
    """
    Sum market_value across all holdings entries.  Because holdings only contains
    securities (cash is never in the holdings array), this always equals the true
    total securities value regardless of how the broker labels subtotals on the
    statement — more reliable than trusting a single extracted field.
    """
    return sum(h.get("market_value") or 0 for h in holdings)

def fix_dynamic_array_formulas(xlsx_bytes: bytes) -> bytes:
    """
    openpyxl strips two things that Excel 365 needs to treat FILTER/SORT/UNIQUE
    as modern dynamic array formulas rather than legacy CSE ({}) formulas:

    1. ca="1" attribute on the <f> element — signals dynamic array spill behaviour.
       openpyxl's ArrayFormula writes t="array" but never adds ca="1".

    2. The xcalcf:calcFeatures extLst block in workbook.xml — without this,
       Excel 365 ignores ca="1" entirely and falls back to {} CSE mode.

    Both are injected here as a post-processing step on the raw xlsx zip bytes,
    since openpyxl has no API to set either.
    """
    # Feature declarations copied verbatim from a file saved by Excel 365
    CALC_FEATURES_EXTLST = (
        '<extLst>'
        '<ext uri="{140A7094-0E35-4892-8432-C4D2E57EDEB5}" '
        'xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main">'
        '<x15:workbookPr chartTrackingRefBase="1"/></ext>'
        '<ext uri="{B58B0392-4F1F-4190-BB64-5DF3571DCE5F}" '
        'xmlns:xcalcf="http://schemas.microsoft.com/office/spreadsheetml/2018/calcfeatures">'
        '<xcalcf:calcFeatures>'
        '<xcalcf:feature name="microsoft.com:RD"/>'
        '<xcalcf:feature name="microsoft.com:Single"/>'
        '<xcalcf:feature name="microsoft.com:FV"/>'
        '<xcalcf:feature name="microsoft.com:CNMTM"/>'
        '<xcalcf:feature name="microsoft.com:LET_WF"/>'
        '<xcalcf:feature name="microsoft.com:LAMBDA_WF"/>'
        '<xcalcf:feature name="microsoft.com:ARRAYTEXT_WF"/>'
        '</xcalcf:calcFeatures></ext>'
        '</extLst>'
    )
    buf_in  = io.BytesIO(xlsx_bytes)
    buf_out = io.BytesIO()
    ca_pattern = re.compile(
        r'<f t="array" ref="([^"]+)">(_xlfn\._xlws\.(?:FILTER|SORT|UNIQUE|SORTBY|SEQUENCE))'
    )
    with zipfile.ZipFile(buf_in, "r") as zin, \
         zipfile.ZipFile(buf_out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            # Fix 1: add ca="1" to dynamic array formulas in each worksheet
            if item.filename.startswith("xl/worksheets/sheet") and b"_xlfn._xlws." in data:
                xml  = data.decode("utf-8")
                xml  = ca_pattern.sub(r'<f t="array" ref="\1" ca="1">\2', xml)
                data = xml.encode("utf-8")
            # Fix 2: inject calcFeatures into workbook.xml so Excel honours ca="1"
            if item.filename == "xl/workbook.xml":
                xml = data.decode("utf-8")
                if "xcalcf:calcFeatures" not in xml:
                    xml = xml.replace("</workbook>", CALC_FEATURES_EXTLST + "</workbook>")
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    return buf_out.getvalue()


# ── Core builder ──────────────────────────────────────────────────────────────
def build_excel(extract: dict, template_bytes: bytes) -> bytes:
    wb = load_workbook(io.BytesIO(template_bytes))

    # Unwrap LlamaExtract's "data" envelope if present
    if "data" in extract and isinstance(extract["data"], dict):
        extract = extract["data"]

    holdings     = deduplicate_holdings(extract.get("holdings", []))
    transactions = extract.get("transactions", [])
    lookup        = build_symbol_lookup(holdings)

    # Normalize all trans_type values
    for t in transactions:
        t["trans_type"] = normalize_trans_type(t.get("trans_type", ""))

    confirmed = [t for t in transactions if not t.get("is_pending")]
    pending   = [t for t in transactions if t.get("is_pending")]

    # ── 1. Transaction Glossary ──────────────────────────────────────────────
    ref = wb["Transaction Glossary"]

    for r in range(5, 100):
        clear_row(ref, r, PEND_COLS)

    conf_start = 5
    for i, t in enumerate(confirmed):
        r = conf_start + i
        set_date(ref[f"D{r}"], parse_date(t.get("date")))
        ref[f"E{r}"].value = t.get("trans_type", "")
        ref[f"F{r}"].value = resolve_symbol(t.get("description"), t.get("symbol"), lookup)
        ref[f"G{r}"].value = t.get("description", "")
        ref[f"H{r}"].value = t.get("currency", "")
        ref[f"I{r}"].value = t.get("price_per_share")
        ref[f"J{r}"].value = t.get("shares")
        ref[f"K{r}"].value = t.get("amount")
        ref[f"L{r}"].value = t.get("charges") or None

    pend_title_row  = conf_start + len(confirmed) + 2
    pend_data_start = pend_title_row + 1
    apply_blue_row(ref, pend_title_row, PEND_COLS,
                   label="Pending Transactions", label_col="C")

    for i, t in enumerate(pending):
        r = pend_data_start + i
        ref[f"C{r}"].value = "Pending"
        set_date(ref[f"D{r}"], parse_date(t.get("date")))
        ref[f"E{r}"].value = t.get("trans_type", "")
        ref[f"F{r}"].value = resolve_symbol(t.get("description"), t.get("symbol"), lookup)
        ref[f"G{r}"].value = t.get("description", "")
        ref[f"H{r}"].value = t.get("currency", "")
        ref[f"I{r}"].value = t.get("price_per_share")
        ref[f"J{r}"].value = t.get("shares")
        set_date(ref[f"N{r}"], parse_date(t.get("settle_date")))
        ref[f"P{r}"].value = t.get("amount")

    # ── 2. Individual stock tabs ─────────────────────────────────────────────
    template_sheet = wb["Individual Stock Template"]

    tickers = sorted(set(
        resolve_symbol(t.get("description",""), t.get("symbol"), lookup)
        for t in transactions
        if t.get("trans_type","").lower() in ("buy","sell")
        and resolve_symbol(t.get("description",""), t.get("symbol"), lookup)
    ))

    for ticker in tickers:
        new_ws = wb.copy_worksheet(template_sheet)
        new_ws.title = ticker[:31]
        new_ws["D2"] = ticker
        # All other formulas (C6 FILTER, R4/R5/R6/R7 outputs) are intentionally
        # NOT overwritten — copy_worksheet() copies them correctly from the template,
        # so any formula updates made directly in Investment_Tracking_V3.xlsx are
        # automatically picked up without needing a main.py change.

    # Re-order sheets alphabetically
    fixed = ["Summary", "Transaction Glossary", "Individual Stock Template"]
    stock_tabs = sorted([s for s in wb.sheetnames if s not in fixed])
    for i, name in enumerate(fixed + stock_tabs):
        wb.move_sheet(name, offset=wb.sheetnames.index(name) - i)

    # ── 3. Summary sheet ─────────────────────────────────────────────────────
    summary = wb["Summary"]

    # Build opening-positions lookup keyed by symbol (upper)
    opening_lookup = {}
    for op in extract.get("opening_positions") or []:
        sym = (op.get("symbol") or "").upper().strip()
        if sym:
            opening_lookup[sym] = op

    sorted_holdings = sorted(holdings, key=lambda x: x.get("symbol") or "")
    last_data_row = 4 + len(sorted_holdings) - 1  # used for reconciliation SUM range

    for i, h in enumerate(sorted_holdings):
        r = 4 + i
        ticker = h.get("symbol") or lookup.get((h.get("description", "")).upper().strip(), "")
        ticker_upper = ticker.upper().strip()

        # ── Section 1: Asset Identifier ──────────────────────────────────────
        summary[f"C{r}"].value = h.get("description", "")
        summary[f"D{r}"].value = ticker
        summary[f"E{r}"].value = h.get("currency", "")
        summary[f"F{r}"].value = extract.get("fx_rate")

        # ── Section 2: Beginning Balances (conditional — only if in opening_positions) ──
        op = opening_lookup.get(ticker_upper)
        if op:
            summary[f"H{r}"].value = op.get("opening_book_cost")
            summary[f"I{r}"].value = op.get("opening_quantity")
            summary[f"J{r}"].value = op.get("opening_acb_per_share")

        # ── Section 3: Net Change in Position (SUMPRODUCT formulas) ──────────
        gl = "'Transaction Glossary'"
        sym_match = f"({gl}!F$5:F$999=D{r})"
        # NOTE: all formulas guarded with IF(D="","",...) — in Excel a blank D
        # matches every blank F cell (empty = empty is TRUE), which made each
        # no-ticker Summary row sum ALL symbol-less transactions.
        blank_guard = f"$D{r}=\"\""
        summary[f"L{r}"].value = (
            f"=IF({blank_guard},\"\",SUMPRODUCT({sym_match}*({gl}!J$5:J$999)))"
        )
        summary[f"M{r}"].value = (
            f"=IF({blank_guard},\"\",SUMPRODUCT({sym_match}*({gl}!K$5:K$999)))"
        )
        summary[f"N{r}"].value = (
            f"=IF({blank_guard},\"\",SUMPRODUCT({sym_match}*({gl}!E$5:E$999=\"Dividend\")*({gl}!K$5:K$999)))"
        )

        # ── Section 4: Ending Balances (formulas; skip T, U, V) ──────────────
        summary[f"P{r}"].value = f"=IF({blank_guard},\"\",H{r}+M{r})"
        summary[f"Q{r}"].value = f"=IF({blank_guard},\"\",I{r}+L{r})"
        summary[f"R{r}"].value = f"=IFERROR(P{r}/Q{r},\"\")"
        summary[f"S{r}"].value = f"=IFERROR(INDIRECT(D{r}&\"!R4\"),\"\")"
        # T (FMV/Unit), U (FMV Total), V (Unrealized Gain) — left for FMV build

    # ── Section 5: Reconciliation block (rows 4–6, fixed positions) ──────────
    fmv_sum_range = f"U4:U{last_data_row}"
    summary["Y4"].value = f"=SUM({fmv_sum_range})"
    summary["Y5"].value = "=Y4"
    summary["Y6"].value = "=Y4+Y5"

    summary["Z4"].value = compute_securities_value(holdings)
    summary["Z5"].value = extract.get("closing_cash_balance")
    summary["Z6"].value = extract.get("total_account_value")

    for row_num in (4, 5, 6):
        summary[f"AA{row_num}"].value = f"=Y{row_num}-Z{row_num}"
        summary[f"AB{row_num}"].value = f'=IF(AA{row_num}=0,"✓","!")'

    output = io.BytesIO()
    wb.save(output)
    return fix_dynamic_array_formulas(output.getvalue())

# ── Routes ────────────────────────────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/trigger")
async def trigger_workflow(file: UploadFile = File(...)):
    """
    Receives a PDF from the UI, forwards it to n8n, decodes the
    base64 Excel response, and returns raw binary to the browser.
    """
    if not N8N_WEBHOOK_URL:
        raise HTTPException(500, "N8N_WEBHOOK_URL environment variable not set")

    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(400, "Only PDF files are accepted")

    pdf_bytes = await file.read()
    xlsx_filename = file.filename.rsplit(".", 1)[0] + ".xlsx"

    async with httpx.AsyncClient(timeout=600.0) as client:
        resp = await client.post(
            N8N_WEBHOOK_URL,
            files={"Investment_Statement": (file.filename, pdf_bytes, "application/pdf")},
        )
        if resp.status_code not in (200, 202):
            raise HTTPException(502, f"n8n webhook returned {resp.status_code}")

    # n8n returns JSON with base64-encoded Excel — decode in Python
    payload = resp.json()
    excel_bytes = base64.b64decode(payload["data"])
    filename = payload.get("filename", xlsx_filename)

    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename}\""}
    )


@app.post("/build-excel")
async def build_excel_endpoint(req: BuildRequest):
    # Fetch template from URL set in env var (upload template to GitHub or S3)
    if not TEMPLATE_URL:
        raise HTTPException(500, "TEMPLATE_URL environment variable not set")

    async with httpx.AsyncClient() as client:
        resp = await client.get(TEMPLATE_URL)
        if resp.status_code != 200:
            raise HTTPException(500, f"Failed to fetch template: {resp.status_code}")
        template_bytes = resp.content

    excel_bytes = build_excel(req.extract_result, template_bytes)
    b64 = base64.b64encode(excel_bytes).decode("utf-8")

    period = req.extract_result.get("statement_period_end", "output")
    broker = req.extract_result.get("broker_name", "Investment").replace(" ", "_")
    filename = f"{broker}_{period}.xlsx"

    return JSONResponse({
        "filename": filename,
        "data": b64,
        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    })
